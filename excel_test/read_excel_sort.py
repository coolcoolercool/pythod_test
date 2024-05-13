# -*- coding: utf-8 -*
# 将 excel 中的内容读取到map中再进行排序

import time
import openpyxl


class SaleContent:
    sale_date = ''
    sale_area = ''
    sale_channel = ''
    sale_order_no = ''
    brand = ''
    sale_price = 0
    sale_num = 0

    # 打印对象的所有属性，类似Java的ToString
    def __str__(self):
        return ' '.join(('%s' % item for item in self.__dict__.values()))


class BrandData:
    order_no_list = []
    count_num = 0
    brand_name = ''

    def __str__(self):
        return ' '.join(('%s' % item for item in self.__dict__.values()))


def read_excel(excel_file_path, sheet_name_input):
    # 打开工作簿
    workbook = openpyxl.load_workbook(excel_file_path)
    # 获取指定表单
    sheet = workbook[sheet_name_input]
    # 获取行数和列数
    print("row_num:" + str(sheet.max_row))
    print("col_num:" + str(sheet.max_column))

    # 创建新的 sheet 或者
    count_content_map = {}

    for row in range(2, sheet.max_row + 1):
        excel_row = SaleContent()
        excel_row.sale_date = sheet.cell(row, 1).value.strftime('%Y/%m/%d/')
        excel_row.sale_area = sheet.cell(row, 2).value
        excel_row.sale_channel = sheet.cell(row, 3).value
        excel_row.sale_order_no = sheet.cell(row, 4).value
        excel_row.brand = sheet.cell(row, 5).value
        excel_row.sale_num = sheet.cell(row, 6).value
        count_content_map[excel_row.sale_order_no] = excel_row

    print("\n关键词频率统计:")
    for value in count_content_map.values():
        pass
        # print(value)

    brand_data_map = {}
    for value in count_content_map.values():
        brand_name = value.brand

        if brand_name not in brand_data_map:
            one_brand_data = BrandData()
            one_brand_data.count_num = 1
            one_brand_data.order_no_list = []
        else:
            one_brand_data = brand_data_map[brand_name]
            one_brand_data.count_num = one_brand_data.count_num + 1
        one_brand_data.brand_name = brand_name
        one_brand_data.order_no_list.append(value.sale_order_no)
        brand_data_map[brand_name] = one_brand_data

    # 针对 count_num 来排序
    for key, value in sorted(brand_data_map.items(), key=lambda kv: kv[1].count_num):
        print(value.brand_name, value.count_num, value.order_no_list, len(value.order_no_list))
        print()


if __name__ == "__main__":
    print("main begin\n")
    now_time = str(int(time.time()))
    # 1.将文件放到data_file目录中

    # 2.输入excel的文件名
    file_name = '2020年销售数据'
    file_path = '../data_file/' + file_name + '.xlsx'
    output_file_path = '../data_file/+ file_name +_' + now_time + '.xlsx'

    # 3.输入你要查询和过滤的sheet 名称
    sheet_name = '销售数据第一页'

    # 4.输入你想过滤和筛选的关键词集合
    set_channel = {'拼多多', '抖音'}

    # 5.输入你想过滤和筛选的关键词在哪里列
    key_word_col_num = 3

    read_excel(file_path, sheet_name)
