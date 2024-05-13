# -*- coding: utf-8 -*
# 将 excel 中的 sheet 简单合并，保存到新的文件中
import time
import openpyxl
import pandas as pd
from pandas import DataFrame


def read_excel(excel_file_path, sheet_names_input, output_file_path):
    # 打开工作簿
    workbook = openpyxl.load_workbook(excel_file_path)

    if sheet_names_input is None:
        sheet_names_input = workbook.sheetnames
    print(sheet_names_input)

    alldata = DataFrame()
    for sheet_name in sheet_names_input:
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None)
        if df.empty:
            continue
        else:
            alldata = pd.concat([alldata, df], axis=0)

    print(alldata)
    alldata.to_excel(output_file_path)


# 这个可以保留根式，更好
def read_excel_v1(file_path_input, sheet_name_input, output_file_path):
    wb = openpyxl.load_workbook(file_path_input)
    if sheet_name_input is not None and len(sheet_name_input) > 1:
        ws = wb[sheet_name_input[0]]
    else:
        sheet_name_input = wb.sheetnames

    for index in range(1, len(sheet_name_input)):
        sheet_x = wb[sheet_name_input[index]]
        for sheet_row in list(sheet_x.values)[1:]:  ## 对表 进行切片，从第2行开始。目的是为了去除表头（可以自行设置）
            ws.append(sheet_row)  # 合并表格
            pass
        pass
    wb.save(output_file_path)


if __name__ == "__main__":
    print("main begin\n")
    now_time = str(int(time.time()))
    # 1.将文件放到data_file目录中

    # 2.输入excel的文件名
    file_name = '2020年销售数据'
    file_path = '../data_file/' + file_name + '.xlsx'
    output_file_path = '../data_file/' + file_name + '_new_' + now_time + '.xlsx'

    # 3.输入想要合并的sheet_name集合，不输入，是将所有的sheet_name 合并
    sheet_name_input = ['销售数据第一页', 'sheet_new_1715150305']

    #read_excel(file_path, sheet_name_input, output_file_path)
    read_excel_v1(file_path, sheet_name_input, output_file_path)
    print('合并完成')
