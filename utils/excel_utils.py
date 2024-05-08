# -*- coding: utf-8 -*
import copy
import datetime
import time

import xlrd
import openpyxl


# 读取excel文件
# xlrd 仅支持 xls 文件格式的excel
def read_excel_file_by_xlrd(excel_file_path):
    # 使用xlrd模块的open_workbook函数打开指定Excel文件并获得Book对象（工作簿）
    wb = xlrd.open_workbook(excel_file_path)
    # 通过Book对象的sheet_names方法可以获取所有表单名称
    sheetnames = wb.sheet_names()
    print(sheetnames)
    # 通过指定的表单名称获取Sheet对象（工作表）
    sheet = wb.sheet_by_name(sheetnames[0])
    # 通过Sheet对象的 nrows 和 ncols 属性获取表单的行数和列数
    print(sheet.nrows, sheet.ncols)
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            # 通过Sheet对象的cell方法获取指定Cell对象（单元格）
            # 通过Cell对象的value属性获取单元格中的值
            value = sheet.cell(row, col).value
            print(value, end='\t')
        print()
    # 获取最后一个单元格的数据类型
    # 0 - 空值，1 - 字符串，2 - 数字，3 - 日期，4 - 布尔，5 - 错误
    last_cell_type = sheet.cell_type(sheet.nrows - 1, sheet.ncols - 1)
    print(last_cell_type)
    # 获取第一行的值（列表）
    print(sheet.row_values(0))
    # 获取指定行指定列范围的数据（列表）
    # 第一个参数代表行索引，第二个和第三个参数代表列的开始（含）和结束（不含）索引
    print(sheet.row_slice(3, 0, 5))


# 读取excel文件
# openpyxl 仅支持 2007年excel版本 文件格式的excel
def read_excel_file_by_openpyxl(excel_file_path):
    # 加载一个工作簿 ---> Workbook
    wb = openpyxl.load_workbook(excel_file_path)
    # 获取工作表的名字
    print(wb.sheetnames)
    # 获取工作表 ---> Worksheet
    sheet = wb.worksheets[0]
    # 获得单元格的范围
    # print(sheet.dimensions)
    # 获得行数和列数
    print(sheet.max_row, sheet.max_column)

    # 获取指定单元格的值
    # print(sheet.cell(3, 3).value)

    # 需要查询关键词合集
    set_channel = {'拼多多', '抖音'}
    count_map = {}  # 统计每个关键词出现多少次
    count_total = 0
    # 读取所有单元格的数据
    print("print excel all data\n")
    for row_ch in range(2, sheet.max_row + 1):
        print(row_ch, end='\t')
        # 请几行包含需要筛选的关键词
        channel_value = sheet.cell(row_ch, 3).value
        print_flag = channel_value in set_channel
        if print_flag:
            count_map[channel_value] = count_map.get(channel_value, 0) + 1
            count_total = count_total + 1

        for col_ch in 'ABCDEFG':
            value = sheet[f'{col_ch}{row_ch}'].value
            # if type(value) == datetime.datetime:
            #     print(value.strftime('%Y年%m月%d日'), end='\t')
            # elif type(value) == int:
            #     print(f'{value:<10d}', end='\t')
            # elif type(value) == float:
            #     print(f'{value:.4f}', end='\t')
            # else:
            if print_flag:
                print(value, end='\t')
        print()

    print(count_map)
    print(count_total)


def read_excel_save_need(excel_file_path, sheet_name_input):
    # 打开工作簿
    workbook = openpyxl.load_workbook(excel_file_path)
    # 获取指定表单
    sheet = workbook[sheet_name_input]
    # 获取行数和列数
    print("row_num:" + str(sheet.max_row))
    print("col_num:" + str(sheet.max_column))

    # 创建新的 sheet 或者
    now_time = str(int(time.time()))
    sheet_name_new = 'sheet_new_' + now_time
    new_sheet = workbook.create_sheet(sheet_name_new)
    # 需要查询关键词合集
    count_map = {}  # 统计每个关键词出现多少次
    count_total = 0
    data = []  # 过滤得到所有需要的数据行
    new_row_num = 2
    # 读取所有单元格的数据
    for row in range(1, sheet.max_row + 1):
        content = []
        # 过滤筛选
        channel_value = sheet.cell(row, key_word_col_num).value
        print_flag = channel_value in set_channel
        if row == 1:
            print_flag = True
        if print_flag and row > 1:
            count_map[channel_value] = count_map.get(channel_value, 0) + 1
            count_total = count_total + 1

        for col in range(1, sheet.max_column + 1):
            source_cell = sheet.cell(row, col)
            if print_flag and source_cell.has_style:
                target_cell = new_sheet.cell(new_row_num, col, source_cell.value)
                # 保留样式复制
                # 设置样式 得用 copy.copy() 不然会报错
                target_cell._style = copy.copy(source_cell._style)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)
            if type(source_cell.value) == datetime.datetime:
                content.append(source_cell.value.strftime('%Y/%m/%d/'))
            else:
                content.append(source_cell.value)
        if print_flag:
            new_row_num += 1
            data.append(content)

    # 打印或者处理过滤筛选后的结果
    for element in data:
        print(element)

    print("\n关键词频率统计:")
    print(count_map)
    print(count_total)

    # 如果文件被占用(比如被wps或者office打开，这里会报错)
    workbook.save(excel_file_path)

def openpyxl_test_function():
    now_time = str(int(time.time()))
    file_name = '../data_file/test' + now_time + '.xlsx'

    # 创建workbook对象
    wb = openpyxl.Workbook()
    # 删除默认创建的一个sheet页
    ws = wb['Sheet']
    wb.remove(ws)

    sheet_name = 'sheet_' + now_time
    ws = wb.create_sheet(sheet_name)
    # 写入方式1：（行、列、值）
    ws.cell(row=1, column=1, value="python")
    ws.cell(row=2, column=1).value = "java"
    ws.cell(row=3, column=1).value = 2
    ws['A4'].value = 3

    # 写入方式2：append方法
    row = ["A11", "A12", "A13"]
    # 向工作表中 按行添加数据
    ws.append(row)

    # 写入方式3：批量写入，for循环方法
    link_nums = 100
    for i in range(2, link_nums):
        ws['A' + str(i)] = "WZ202207FF" + str(i)
        ws['B' + str(i)] = "147258369"
        ws['C' + str(i)] = "DDYY"

    # 保存xlsx
    wb.save(file_name)


class CasesData:
    """用于保存测试用例数据"""
    pass


class ReadExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sh = self.wb[self.sheet_name]

    def close(self):
        self.wb.close()

    def read_data(self):
        """按行读取数据，最后返回一个存储字典的列表"""
        self.open()
        rows = list(self.sh.rows)
        titles = []
        for t in rows[0]:
            title = t.value
            titles.append(title)
        cases = []
        for row in rows[1:]:
            case = []
            for r in row:
                case.append(r.value)
            cases.append(dict(zip(titles, case)))
        self.close()
        return cases

    def write_data(self, row, column, msg):
        self.open()
        self.sh.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)
        self.close()


if __name__ == "__main__":
    print("main begin\n")
    now_time = str(int(time.time()))
    # 1.将文件放到data_file目录中

    # 2.输入excel的文件名
    file_name = '2020年销售数据'
    file_path = '../data_file/' + file_name + '.xlsx'
    output_file_path = '../data_file/+ file_name +_' + now_time + '.xlsx'

    # 3.输入你要查询和过滤的sheet 名称
    sheet_name = '销售数据第一页'

    # 4.输入你想过滤和筛选的关键词集合
    set_channel = {'拼多多', '抖音'}

    # 5.输入你想过滤和筛选的关键词在哪里列
    key_word_col_num = 3

    read_excel_save_need(file_path, sheet_name)

    # 直接运行本文件时执行，下面是一个应用实例
    # 需要读取excel时直接调用ReadExcel类
    # test2 = ReadExcel('../data_file/2020年销售数据.xlsx', '销售数据第一页')
    # res2 = test2.read_data()	# 最后返回一个存储字典的列表
    # print(res2[0]['销售日期'])	# 打印第1个用例的case_id
    # print(res2[3]['售价'])		# 打印第4个用例的title
